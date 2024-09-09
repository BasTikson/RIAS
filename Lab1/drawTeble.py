from operator import index

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import subprocess
from datetime import datetime
from constant import obs_ir


# ToDo Центрировать шапку таблицы
# ToDo Сохранение отчетов в папках

# ToDo Подумать на каком этапе и в каком формате будут проводиться расчеты
# Водим еще одну функцию, которая будет просчитывать и заполнять массивы типа [number_IR, "Приведенаая стоимость приобретения", "", "", "", "", ""]
# Вызов класса происходит, только после расчета ИР
# Подумать и переорганизовать функции в lab_1 в полноценный класс

years_list = [2018, 2019, 2020, 2021, 2022]

class ExcelGenerator:
    def __init__(self, obs_ir_info:dict[int], years_list_info:list[int]):

        self.name_file = f"Отчет {datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
        self.output_path_pdf = "pdf/" + self.name_file
        self.output_path_xlsx = "xlsx/" + self.name_file



        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Оценки стоимости ИР 1 категории"


        self.headers = [ "№ИР", "Показатель", "Значение показателя в год, тыс. руб"] + ([""] * (len(years_list_info) - 1))
        self.years = ["", ""] + years_list_info
        self.data = []
        self.index_merge_cells = {}
        self.obs_ir = obs_ir_info
        self.build_table(obs_ir_info)

    def build_table(self, obs_ir_info, index_merge_cells_stop:int = 2):
        index_merge_cells_start = index_merge_cells_stop

        for i in obs_ir_info.items():
            list_row_property_IR = []
            number_IR = i[0]
            property_IR = i[1][0]

            if "приобретаемый" in property_IR:
                list_row_property_IR.append([number_IR, "Стоимость приобретения", "", "", "", "", ""])
                list_row_property_IR.append([number_IR, "Приведенаая стоимость приобретения", "", "", "", "", ""])
                index_merge_cells_stop = index_merge_cells_stop + 2

            if "разрабатываемый" in property_IR:
                list_row_property_IR.append([number_IR, "Базовая стоимость разработки", "", "", "", "", ""])
                list_row_property_IR.append([number_IR, "Накопительная стоимость разработки", "", "", "", "", ""])
                # list_row_property_IR.append([number_IR, "Приведённая стоимость эксплуатации", "", "", "", "", ""])
                index_merge_cells_stop = index_merge_cells_stop + 2

            if "обслуживаемый" in property_IR:
                list_row_property_IR.append([number_IR, "Стоимость обслуживания", "", "", "", "", ""])
                index_merge_cells_stop = index_merge_cells_stop + 1

            if "приносящий прибыль" in property_IR:
                list_row_property_IR.append([number_IR, "Прибыль", "", "", "", "", ""])
                index_merge_cells_stop = index_merge_cells_stop + 1

            list_row_property_IR.append([number_IR, "ОБЩАЯ стоимость", "", "", "", "", ""])
            index_merge_cells_stop = index_merge_cells_stop + 1
            self.index_merge_cells.update({index_merge_cells_start + 1: index_merge_cells_stop})
            index_merge_cells_start = index_merge_cells_stop

            self.data = self.data + list_row_property_IR

    def generate_excel(self):
        # Установка значений в первую ячейку диапазона
        self.ws['A1'] = self.headers[0]
        self.ws['B1'] = self.headers[1]
        self.ws['C1'] = self.headers[2]

        # Установка значений для годов
        for col_num, year in enumerate(self.years, 1):
            cell = self.ws.cell(row=2, column=col_num)
            cell.value = year
            cell.font = Font(bold=True)

        # Установка значений в первую ячейку диапазона
        self.ws['A3'] = self.data[0][0]

        for row_num, row_data in enumerate(self.data, 3):
            for col_num, cell_data in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num)
                cell.value = cell_data

        # Преобразуем генератор в список
        columns = list(self.ws.columns)

        # Автоматически устанавливаем ширину столбцов до столбца C
        for column_cells in columns[:2]:  # Столбцы до C (индекс 3)
            length = max(len(str(cell.value)) for cell in column_cells)
            self.ws.column_dimensions[
                column_cells[0].column_letter].width = length + 2  # Добавляем немного дополнительного пространства

        # Объединение ячеек для данных
        # №ИР
        self.ws.merge_cells('A1:A2')
        # Показатель
        self.ws.merge_cells('B1:B2')
        # Значение показателя в год, тыс. руб
        self.ws.merge_cells('C1:G1')
        # Индексы №ИР
        for i in self.index_merge_cells.items():
            self.ws.merge_cells(f'A{i[0]}:A{i[1]}')

        # Определяем стиль границы
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Применяем границы ко всем ячейкам
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=self.ws.max_column):
            for cell in row:
                cell.border = thin_border

        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Сохранение Excel файла
        excel_file = 'example.xlsx'
        self.wb.save(excel_file)

        return excel_file

    def convert_to_pdf(self, excel_file):
        # Конвертация Excel в PDF с использованием libreoffice
        pdf_file = 'example.pdf'
        subprocess.run(['libreoffice', '--headless', '--convert-to', 'pdf', excel_file, '--outdir', '.'])
        return pdf_file

    def open_pdf(self, pdf_file):
        # Открытие PDF файла
        subprocess.Popen(['xdg-open', pdf_file], shell=False)

    def run(self):
        excel_file = self.generate_excel()
        pdf_file = self.convert_to_pdf(excel_file)
        self.open_pdf(pdf_file)


generator = ExcelGenerator(obs_ir, years_list)
generator.run()